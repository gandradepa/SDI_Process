# File: SDI_Spreadsheet.py
# Purpose: Load from SQLite -> df (keep only required cols) -> df2 (renamed + constants)
#          -> Fill Excel template -> Mark rows as exported in DB.

import os
import re
import sqlite3
from datetime import datetime
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

# --------------------------------------------------------------------------------------
# CONFIG
# --------------------------------------------------------------------------------------
DB_PATH = r"S:\MaintOpsPlan\AssetMgt\Asset Management Process\Database\8. New Assets\Git_control\SDI Process\QR_codes.db"
TEMPLATE_PATH = r"S:\MaintOpsPlan\AssetMgt\Asset Management Process\Database\8. New Assets\Git_control\SDI Process\Import Assets-TEMPLATE-082923.xlsx"
OUTPUT_DIR = r"S:\MaintOpsPlan\AssetMgt\Asset Management Process\Database\8. New Assets\Git_control\SDI Process"

# Excel template layout
HEADER_ROW = 9   # template headers row
START_ROW = 10   # first data row

# Output naming
FILE_PREFIX = "SDI_Process"
DATE_FMT = "%m_%d_%Y"

# Export behavior
GROUP_BY_PROPERTY = False  # False => single workbook containing all rows. True => one file per Property

# --------------------------------------------------------------------------------------
# SDI source columns to keep
# --------------------------------------------------------------------------------------
KEEP_COLS: List[str] = [
    "QR Code", "Building", "Description", "Asset Group", "UBC Tag", "Serial", "Model",
    "Manufacturer", "Attribute", "Ampere", "Supply From", "Volts", "Location",
    "Diameter", "Technical Safety BC", "Year",
]

# Rename -> df2 target schema (as per screenshot)
COLUMN_RENAME_MAP: Dict[str, str] = {
    "QR Code": "Code",
    "Building": "Property",
    "Description": "Description",
    "Asset Group": "Asset Group",
    "UBC Tag": "Asset Tag",
    "Serial": "Serial Number",
    "Model": "Model",
    "Manufacturer": "Make",
    "Attribute": "Attribute Set",
    "Ampere": "Amperage Rating",
    "Supply From": "Fed From Equipment ID",
    "Volts": "Voltage Rating",
    "Location": "Space Details",
    "Diameter": "Diameter",
    "Technical Safety BC": "Previous (OLD) ID",
    "Year": "Date Of Manufacture Or Construction",
}
NEW_COLS: List[str] = [COLUMN_RENAME_MAP.get(c, c) for c in KEEP_COLS]

# Constant columns required by template
CONST_COLS: Dict[str, object] = {
    "Is Missing (Y/N)": False,
    "Simple": True,
    "Is Planned Maintenance Required? (Y/N)": False,
}
FINAL_COLS: List[str] = NEW_COLS + list(CONST_COLS.keys())

# --------------------------------------------------------------------------------------
# Load -> df (only KEEP_COLS)
# --------------------------------------------------------------------------------------
def load_sdi_print_out(db_path: str = DB_PATH) -> pd.DataFrame:
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database not found at: {db_path}")

    with sqlite3.connect(db_path, timeout=15) as conn:
        cols_sql = ", ".join([f'"{c}"' for c in KEEP_COLS])
        try:
            df = pd.read_sql_query(f"SELECT {cols_sql} FROM sdi_print_out", conn)
        except Exception:
            df = pd.read_sql_query("SELECT * FROM sdi_print_out", conn)

    for c in KEEP_COLS:
        if c not in df.columns:
            df[c] = ""
    return df.loc[:, KEEP_COLS]

# --------------------------------------------------------------------------------------
# Build df2 (rename + constants)
# --------------------------------------------------------------------------------------
def rename_for_df2(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.rename(columns=COLUMN_RENAME_MAP)
    for col in NEW_COLS:
        if col not in df2.columns:
            df2[col] = ""
    for name, value in CONST_COLS.items():
        df2[name] = value
    return df2.loc[:, FINAL_COLS]

# --------------------------------------------------------------------------------------
# Excel helpers
# --------------------------------------------------------------------------------------
def _normalize_name(text: str) -> str:
    s = "" if text is None else str(text)
    s = re.sub(r"[^0-9a-zA-Z]+", " ", s).strip().lower()
    return re.sub(r"\s+", " ", s)

def _safe_filename(text: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", str(text)).strip()

def _unique_path(path: str) -> str:
    base, ext = os.path.splitext(path)
    i, candidate = 1, path
    while os.path.exists(candidate):
        candidate = f"{base} ({i}){ext}"
        i += 1
    return candidate

def _build_output_path(property_label: str) -> str:
    date_s = datetime.now().strftime(DATE_FMT)
    prop_s = _safe_filename(property_label or "UnknownProperty")
    return os.path.join(OUTPUT_DIR, f"{FILE_PREFIX}_{date_s}_{prop_s}.xlsx")

def _map_headers_to_df2_cols(ws, df2: pd.DataFrame) -> Dict[int, str]:
    norm_df_cols = {_normalize_name(c): c for c in df2.columns}
    mapping: Dict[int, str] = {}
    for col_idx in range(1, ws.max_column + 1):
        header_val = ws.cell(row=HEADER_ROW, column=col_idx).value
        norm_header = _normalize_name(header_val)
        if norm_header in norm_df_cols:
            mapping[col_idx] = norm_df_cols[norm_header]
    if not mapping:
        raise ValueError("No template headers (row 9) matched df2 columns.")
    return mapping

def _write_group_to_sheet(ws, mapping: Dict[int, str], df_group: pd.DataFrame) -> None:
    df_group = df_group.reset_index(drop=True)
    for r, (_, row) in enumerate(df_group.iterrows(), start=START_ROW):
        for col_idx, df_col in mapping.items():
            val = row.get(df_col)
            ws.cell(row=r, column=col_idx, value=(None if pd.isna(val) else val))

# --------------------------------------------------------------------------------------
# Export
# --------------------------------------------------------------------------------------
def export_df2_to_template(df2: pd.DataFrame, group_by_property: bool = GROUP_BY_PROPERTY) -> List[str]:
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    prop_col = "Property"
    paths: List[str] = []

    def _single_label_for(df2: pd.DataFrame) -> str:
        if prop_col not in df2.columns:
            return "UnknownProperty"
        uniq = [str(v).strip() for v in df2[prop_col].fillna("").astype(str).unique()]
        uniq = [u for u in uniq if u]
        return uniq[0] if len(uniq) == 1 else "MULTI"

    # Mode A: one workbook with ALL rows
    if not group_by_property:
        label = _single_label_for(df2)
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        mapping = _map_headers_to_df2_cols(ws, df2)
        _write_group_to_sheet(ws, mapping, df2)
        out_path = _unique_path(_build_output_path(label))
        wb.save(out_path)
        paths.append(out_path)
        return paths

    # Mode B: one workbook per Property
    if prop_col in df2.columns:
        groups: List[Tuple[str, pd.DataFrame]] = []
        for prop_val, grp in df2.groupby(prop_col, dropna=False):
            prop_name = str(prop_val).strip() if pd.notna(prop_val) and str(prop_val).strip() else "UnknownProperty"
            groups.append((prop_name, grp))
    else:
        groups = [("UnknownProperty", df2)]

    for prop_name, grp in groups:
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        mapping = _map_headers_to_df2_cols(ws, df2)
        _write_group_to_sheet(ws, mapping, grp)
        out_path = _unique_path(_build_output_path(prop_name))
        wb.save(out_path)
        paths.append(out_path)

    return paths

# --------------------------------------------------------------------------------------
# Database Update
# --------------------------------------------------------------------------------------
def mark_all_rows_as_printed(db_path: str = DB_PATH):
    """Updates the 'print_out' column to 1 for all rows in sdi_print_out."""
    try:
        with sqlite3.connect(db_path, timeout=15) as conn:
            cur = conn.cursor()
            cur.execute('UPDATE sdi_print_out SET print_out = 1')
            updated_count = cur.rowcount
            conn.commit()
            print(f"[DB Update] Marked {updated_count} rows as printed.")
    except Exception as e:
        print(f"[DB Update ERROR] Could not update rows: {e}")
        raise

# --------------------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------------------
if __name__ == "__main__":
    try:
        df = load_sdi_print_out()
        
        if df.empty:
            print("[OK] No new assets to process.")
        else:
            df2 = rename_for_df2(df)
            print(f"[OK] rows -> df={len(df)}, df2={len(df2)}")  # should be equal
            
            saved = export_df2_to_template(df2, group_by_property=GROUP_BY_PROPERTY)
            print("Saved files:")
            for p in saved:
                print(" -", p)
            
            # After successful export, update the database
            mark_all_rows_as_printed()

    except Exception as e:
        print(f"[ERROR] {e}")
