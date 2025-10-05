import os
import re
import sqlite3
from datetime import datetime
from io import BytesIO
from typing import Dict, List

import pandas as pd
from flask import Flask, render_template, redirect, url_for, flash, request, send_file
from openpyxl import load_workbook

# -----------------------------------------------------------------------------
# Paths
# -----------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR = os.path.join(BASE_DIR, "template")
STATIC_DIR = os.path.join(BASE_DIR, "static")
DB_PATH = r"/home/developer/asset_capture_app_dev/data/QR_codes.db"
TEMPLATE_PATH = r"/home/developer/SDI_process/template/Import Assets-TEMPLATE-082923.xlsx"

LOGO_MAIN_NAME = "ubc_logo.jpg"
LOGO_FAC_NAME = "ubc-facilities_logo.jpg"

# -----------------------------------------------------------------------------
# Flask
# -----------------------------------------------------------------------------
app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR, static_url_path="/static")
app.secret_key = "replace-with-a-strong-secret"

# -----------------------------------------------------------------------------
# Columns & Mappings
# -----------------------------------------------------------------------------
MASTER_COLS = [
    "id_print_out", "QR Code", "Building", "Description", "Asset Group", "UBC Tag", "Serial", "Model",
    "Manufacturer", "Attribute", "Ampere", "Supply From", "Volts", "Location", "Space",
    "Diameter", "Technical Safety BC", "Year"
]

PRINT_OUT_COLS = MASTER_COLS + ["print_out", "date", "time"]

COLUMN_RENAME_MAP: Dict[str, str] = {
    "QR Code": "Code", "Building": "Property", "Description": "Description",
    "Asset Group": "Asset Group", "UBC Tag": "Asset Tag", "Serial": "Serial Number",
    "Model": "Model", "Manufacturer": "Make", "Attribute": "Attribute Set",
    "Ampere": "Amperage Rating", "Supply From": "Fed From Equipment ID",
    "Volts": "Voltage Rating", "Location": "Space Details", "Space": "Space.Space number",
    "Diameter": "Diameter", "Technical Safety BC": "Previous (OLD) ID",
    "Year": "Date Of Manufacture Or Construction",
}

CONST_COLS: Dict[str, object] = {
    "Is Missing (Y/N)": False, "Simple": True, "Is Planned Maintenance Required? (Y/N)": False,
}

# -----------------------------------------------------------------------------
# Helpers
# -----------------------------------------------------------------------------
def table_exists(conn, table_name):
    cur = conn.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
    return cur.fetchone() is not None

def ensure_columns_and_order(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df = df.loc[:,~df.columns.duplicated()]
    for c in MASTER_COLS:
        if c not in df.columns:
            df[c] = ""
    return df[MASTER_COLS]

def filter_approved(df: pd.DataFrame) -> pd.DataFrame:
    if "Approved" not in df.columns:
        return df.copy()
    return df.loc[df["Approved"].astype(str) == "1"].copy()

def build_sdi_dataset(building_code: str = None) -> pd.DataFrame:
    try:
        if not os.path.exists(DB_PATH):
            raise FileNotFoundError(f"Database not found at: {DB_PATH}")
        with sqlite3.connect(DB_PATH, timeout=10) as conn:
            me = pd.read_sql_query("SELECT * FROM sdi_dataset;", conn)
            el = pd.read_sql_query("SELECT * FROM sdi_dataset_EL;", conn)

        me, el = filter_approved(me), filter_approved(el)
        
        if building_code:
            me = me[me['Building'].astype(str) == str(building_code)]
            el = el[el['Building'].astype(str) == str(building_code)]

        if "UBC Asset Tag" in el.columns and "UBC Tag" not in el.columns:
            el = el.rename(columns={"UBC Asset Tag": "UBC Tag"})

        return pd.concat([me, el], ignore_index=True)
    except Exception as e:
        print(f"[ERROR] in build_sdi_dataset: {repr(e)}")
        raise

def get_codes_in_print_out_table() -> set:
    try:
        with sqlite3.connect(DB_PATH, timeout=10) as conn:
            if not table_exists(conn, "sdi_print_out"):
                return set()
            df_exp = pd.read_sql_query('SELECT DISTINCT "QR Code" FROM sdi_print_out', conn)
        return set(df_exp["QR Code"].astype(str).str.strip().tolist())
    except Exception as e:
        print(f"[ERROR] in get_codes_in_print_out_table: Could not read from sdi_print_out table: {repr(e)}")
        return set()

def get_all_buildings() -> list:
    try:
        with sqlite3.connect(DB_PATH, timeout=10) as conn:
            if not table_exists(conn, 'Buildings'):
                df1 = pd.read_sql_query('SELECT DISTINCT Building FROM sdi_dataset', conn)
                df2 = pd.read_sql_query('SELECT DISTINCT Building FROM sdi_dataset_EL', conn)
                all_codes = sorted(pd.concat([df1, df2])['Building'].dropna().unique())
                return [{'Code': code, 'Name': f'Building {code}'} for code in all_codes]

            df1 = pd.read_sql_query('SELECT DISTINCT Building FROM sdi_dataset', conn)
            df2 = pd.read_sql_query('SELECT DISTINCT Building FROM sdi_dataset_EL', conn)
            asset_building_codes = set(pd.concat([df1, df2])['Building'].dropna().astype(str))

            df_buildings = pd.read_sql_query('SELECT Code, Name FROM Buildings', conn)
            df_buildings['Code'] = df_buildings['Code'].astype(str)
            df_filtered = df_buildings[df_buildings['Code'].isin(asset_building_codes)]
            
            return df_filtered.sort_values('Name').to_dict(orient="records")
    except Exception as e:
        error_msg = f"Could not generate building list: {repr(e)}"
        print(f"[ERROR] in get_all_buildings: {error_msg}")
        flash(f"⚠️ {error_msg}", "danger")
        return []

def build_unpackaged_dataset(building_code: str = None) -> pd.DataFrame:
    try:
        df = build_sdi_dataset(building_code=building_code).copy()
        if df.empty:
            return pd.DataFrame()
            
        df["QR Code"] = df["QR Code"].astype(str).str.strip()
        
        codes_in_print_out = get_codes_in_print_out_table()
        if codes_in_print_out:
            df = df[~df["QR Code"].isin(codes_in_print_out)].copy()

        try:
            with sqlite3.connect(DB_PATH, timeout=10) as conn:
                qr_codes_df = pd.read_sql_query('SELECT "QR_code_ID", "Location" FROM QR_codes', conn)

            qr_codes_df = qr_codes_df.rename(columns={"Location": "Space"})
            
            df['merge_key'] = pd.to_numeric(df['QR Code'], errors='coerce')
            qr_codes_df['merge_key'] = pd.to_numeric(qr_codes_df['QR_code_ID'], errors='coerce')
            
            df = pd.merge(df, qr_codes_df[['merge_key', 'Space']], on='merge_key', how='left')
            df['Space'] = df['Space'].fillna('').astype(str).apply(lambda x: x.split(' ')[0])
            df = df.drop(columns=['merge_key'])

        except Exception as e:
            print(f"[ERROR] in build_unpackaged_dataset (merging data): {repr(e)}")

        return df
    except Exception as e:
        print(f"[ERROR] in build_unpackaged_dataset (main block): {repr(e)}")
        return pd.DataFrame()

def build_packaged_dataset(building_code: str = None) -> pd.DataFrame:
    try:
        with sqlite3.connect(DB_PATH, timeout=10) as conn:
            if not table_exists(conn, 'sdi_print_out'):
                return pd.DataFrame()
            df = pd.read_sql_query('SELECT * FROM sdi_print_out', conn)
            
            if building_code:
                building_name = None
                if table_exists(conn, 'Buildings'):
                    df_buildings = pd.read_sql_query('SELECT Code, Name FROM Buildings', conn)
                    df_buildings['Code'] = df_buildings['Code'].astype(str)
                    match = df_buildings[df_buildings['Code'] == str(building_code)]
                    if not match.empty:
                        building_name = match['Name'].iloc[0]

                code_mask = (df['Building'].astype(str) == str(building_code))
                if building_name:
                    name_mask = (df['Building'].astype(str) == str(building_name))
                    df = df[code_mask | name_mask].copy()
                else:
                    df = df[code_mask].copy()
        
        return df
    except Exception as e:
        print(f"[ERROR] in build_packaged_dataset: {repr(e)}")
        return pd.DataFrame()

def _check_db_writable(path: str):
    folder = os.path.dirname(path) or "."
    if not os.access(folder, os.W_OK):
        raise PermissionError(f"Folder not writable: {folder}")
    if os.path.exists(path) and not os.access(path, os.W_OK):
        raise PermissionError(f"Database file is read-only: {path}")

def _safe_filename(text: str) -> str:
    s = "" if text is None else str(text)
    s = re.sub(r'[\\/:*?"<>|]', "_", s)
    return s.strip()

def _get_building_label_for_filename(df: pd.DataFrame) -> str:
    if "Building" not in df.columns or df.empty:
        return "UnknownBuilding"
    
    uniq = [str(v).strip() for v in df["Building"].fillna("").astype(str).unique()]
    uniq = [u for u in uniq if u]

    if not uniq:
        return "UnknownBuilding"
    elif len(uniq) == 1:
        return _safe_filename(uniq[0])
    else:
        return "MULTI_Building"

def _normalize_name(text: str) -> str:
    s = "" if text is None else str(text)
    s = re.sub(r"[^0-9a-zA-Z]+", " ", s).strip().lower()
    return re.sub(r"\s+", " ", s)

def get_next_sdi_package_id(conn) -> str:
    cur = conn.cursor()
    
    if not table_exists(conn, "sdi_sequence"):
        cur.execute("CREATE TABLE sdi_sequence (last_value INTEGER)")
        initial_value = 0
        try:
            if table_exists(conn, "sdi_print_out"):
                cur.execute('SELECT MAX(id_print_out) FROM sdi_print_out WHERE id_print_out IS NOT NULL AND id_print_out != ""')
                max_id = cur.fetchone()[0]
                if max_id and max_id.startswith("SDI-"):
                    initial_value = int(max_id.split('-')[-1])
        except (sqlite3.OperationalError, IndexError, ValueError):
            pass
        
        cur.execute("INSERT INTO sdi_sequence (last_value) VALUES (?)", (initial_value,))

    cur.execute("SELECT last_value FROM sdi_sequence")
    last_value = cur.fetchone()[0]
    
    new_value = last_value + 1
    
    cur.execute("UPDATE sdi_sequence SET last_value = ?", (new_value,))

    return f"SDI-{new_value:05d}"

# -----------------------------------------------------------------------------
# Routes
# -----------------------------------------------------------------------------
@app.route("/")
def dashboard():
    try:
        selected_building_code = request.args.get("building_code", "")
        
        all_buildings = get_all_buildings()
        unpackaged_df = build_unpackaged_dataset(building_code=selected_building_code)
        packaged_df = build_packaged_dataset(building_code=selected_building_code)

        # Ensure data is formatted correctly for display
        packaged_df = ensure_columns_and_order(packaged_df)
        unpackaged_df = ensure_columns_and_order(unpackaged_df)
        
        sdi_print_controls = []
        if not packaged_df.empty and "id_print_out" in packaged_df.columns:
            sdi_print_controls = sorted(packaged_df[packaged_df["id_print_out"].notna()]["id_print_out"].unique())

        try:
            with sqlite3.connect(DB_PATH, timeout=10) as conn:
                if table_exists(conn, 'Buildings'):
                    df_buildings = pd.read_sql_query('SELECT Code, Name FROM Buildings', conn)
                    df_buildings['Code'] = df_buildings['Code'].astype(str)
                    
                    building_map = pd.Series(df_buildings.Name.values, index=df_buildings.Code).to_dict()

                    if not unpackaged_df.empty:
                        unpackaged_df['Building'] = unpackaged_df['Building'].map(building_map).fillna(unpackaged_df['Building'])

                    if not packaged_df.empty:
                        packaged_df['Building'] = packaged_df['Building'].map(building_map).fillna(packaged_df['Building'])
        except Exception as e:
             print(f"[ERROR] in dashboard (enriching building name): {repr(e)}")
             flash("Could not display building names correctly.", "warning")

        display_rename_map = {"id_print_out": "SDI Print Control"}
        display_columns = [display_rename_map.get(c, c) for c in MASTER_COLS]
        unpackaged_df.rename(columns=display_rename_map, inplace=True)
        packaged_df.rename(columns=display_rename_map, inplace=True)

        unpackaged_df = unpackaged_df.fillna('')
        packaged_df = packaged_df.fillna('')
        
        return render_template(
            "dashboard.html",
            title="SDI - Planon Process Management",
            columns=display_columns,
            unpackaged_rows=unpackaged_df.to_dict(orient="records"),
            packaged_rows=packaged_df.to_dict(orient="records"),
            logo_main_name=LOGO_MAIN_NAME,
            logo_fac_name=LOGO_FAC_NAME,
            all_buildings=all_buildings,
            selected_building=selected_building_code,
            sdi_print_controls=sdi_print_controls
        )
    except Exception as e:
        print(f"[FATAL ERROR] in dashboard route: {repr(e)}")
        flash("A critical error occurred while loading the dashboard. Please check the console log.", "danger")
        return render_template("dashboard.html", title="Error", columns=MASTER_COLS, unpackaged_rows=[], packaged_rows=[], all_buildings=[])

@app.route("/export", methods=["POST"])
def export_to_sdi():
    building_code = request.form.get("building_code")
    force_replace = request.form.get("force_replace", "false").lower() == "true"
    # --- ALTERAÇÃO AQUI ---
    active_tab_anchor = request.form.get("active_tab")

    if not building_code:
        flash("To create a pack, select only one building at time", "warning")
        return redirect(url_for("dashboard", _anchor=active_tab_anchor))

    try:
        df = build_unpackaged_dataset(building_code=building_code) 
        if df.empty:
            flash(f"No new assets to export for the selected building.", "info")
            return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

        required_cols = ["Description", "Asset Group", "Attribute"]
        for col in required_cols:
            if df[col].isnull().any() or df[col].astype(str).str.strip().eq('').any():
                flash('To create a package, the fields "Description", "Asset Group" and "Attribute" must be filled in', "danger")
                return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

        if not force_replace:
            existing_codes = get_codes_in_print_out_table()
            new_codes = set(df["QR Code"].astype(str).str.strip())
            duplicate_codes = list(new_codes.intersection(existing_codes))

            if duplicate_codes:
                message = f"CONFIRM:{','.join(duplicate_codes)}"
                flash(message, "confirmation")
                return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))
        
        _check_db_writable(DB_PATH)
        with sqlite3.connect(DB_PATH, timeout=20) as conn:
            cur = conn.cursor()
            
            conn.execute(f'''CREATE TABLE IF NOT EXISTS sdi_print_out ({", ".join(f'"{col}" TEXT' for col in PRINT_OUT_COLS)})''')
            
            cur.execute("PRAGMA table_info(sdi_print_out)")
            existing_cols = {info[1] for info in cur.fetchall()}
            if "id_print_out" not in existing_cols:
                cur.execute('ALTER TABLE sdi_print_out ADD COLUMN "id_print_out" TEXT')

            new_package_id = get_next_sdi_package_id(conn)
            
            now = datetime.now()
            df_print = df.copy()
            for c in PRINT_OUT_COLS:
                if c not in df_print.columns:
                    df_print[c] = ""

            df_print["id_print_out"] = new_package_id
            df_print["print_out"] = 0
            df_print["date"] = now.strftime("%Y-%m-%d")
            df_print["time"] = now.strftime("%H:%M:%S")
            df_print = df_print.loc[:, PRINT_OUT_COLS]

            if force_replace:
                codes_to_replace = df_print["QR Code"].tolist()
                if codes_to_replace:
                    placeholders = ','.join('?' for _ in codes_to_replace)
                    cur.execute(f'DELETE FROM sdi_print_out WHERE "QR Code" IN ({placeholders})', codes_to_replace)

            df_print.to_sql("sdi_print_out", conn, if_exists="append", index=False)
        
        if force_replace:
            flash(f"✅ Replaced and exported {len(df_print)} rows to package {new_package_id} successfully.", "success")
        else:
            flash(f"✅ Exported {len(df_print)} rows to package {new_package_id} successfully.", "success")

    except Exception as e:
        print(f"[ERROR] in export_to_sdi: {repr(e)}")
        flash(f"⚠️ Could not record the export. {str(e)}", "danger")
    
    return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

@app.route("/exclude_package", methods=["POST"])
def exclude_package():
    sdi_control_id = request.form.get("sdi_control_id")
    building_code = request.form.get("building_code")
    # --- ALTERAÇÃO AQUI ---
    active_tab_anchor = request.form.get("active_tab")

    if not sdi_control_id:
        flash("⚠️ Please select a package to exclude.", "warning")
        return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

    try:
        _check_db_writable(DB_PATH)
        with sqlite3.connect(DB_PATH, timeout=20) as conn:
            cur = conn.cursor()
            cur.execute('DELETE FROM sdi_print_out WHERE "id_print_out" = ?', (sdi_control_id,))
            deleted_rows = cur.rowcount
            conn.commit()
        
        if deleted_rows > 0:
            flash(f"✅ Package {sdi_control_id} ({deleted_rows} assets) has been excluded and returned to Unpackaged Assets.", "success")
        else:
            flash(f"🤔 Package {sdi_control_id} was not found or was already empty.", "info")

    except Exception as e:
        print(f"[ERROR] in exclude_package: {repr(e)}")
        flash(f"⚠️ Could not exclude the package. {str(e)}", "danger")

    return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))


@app.route("/export-planon", methods=["POST"])
def export_to_planon():
    building_code = request.form.get("building_code")
    sdi_control_id = request.form.get("sdi_control_id")
    force_export = request.form.get("force_planon_export", "false").lower() == "true"
    # --- ALTERAÇÃO AQUI ---
    active_tab_anchor = request.form.get("active_tab")
    
    try:
        if not sdi_control_id:
            flash("To export, you must select a unique 'SDI Print Control' value.", "warning")
            return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

        df = build_packaged_dataset(building_code=building_code)
        
        df = df[df["id_print_out"] == sdi_control_id].copy()

        if df.empty:
            flash(f"No assets found for SDI Print Control '{sdi_control_id}'.", "info")
            return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

        with sqlite3.connect(DB_PATH, timeout=15) as conn:
            df_asset_group = pd.DataFrame()
            if table_exists(conn, 'Asset_Group'):
                df_asset_group = pd.read_sql_query('SELECT Name, "Full Classification" FROM Asset_Group', conn)
        
        if not force_export:
            already_exported = df[df["print_out"].astype(str) == "1"]
            if not already_exported.empty:
                codes = already_exported["QR Code"].tolist()
                message = f"PLANON_CONFIRM:{','.join(codes)}"
                flash(message, "planon_confirmation")
                return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

        df_to_export = df[df["print_out"].astype(str) == "0"] if not force_export else df
        if df_to_export.empty and not force_export:
             flash("All assets for this package have already been exported to Planon.", "info")
             return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

        if not df_asset_group.empty:
            panels_mask = df_to_export['Asset Group'].str.strip().str.lower() == 'panels'
            df_to_export.loc[panels_mask, 'Asset Group'] = 'EL.21.306.4067'

            other_assets_mask = ~panels_mask
            asset_groups_to_check = df_to_export.loc[other_assets_mask, 'Asset Group'].str.strip().unique()

            if len(asset_groups_to_check) > 0:
                relevant_mappings = df_asset_group[df_asset_group['Name'].str.strip().isin(asset_groups_to_check)]
                duplicated_names = relevant_mappings[relevant_mappings['Name'].duplicated()]['Name'].unique()

                if duplicated_names.any():
                    conflicting_assets = df_to_export[df_to_export['Asset Group'].isin(duplicated_names)]
                    conflicting_qr_codes = conflicting_assets['QR Code'].tolist()
                    qr_codes_str = ", ".join(conflicting_qr_codes)
                    
                    error_message = f"The Asset Group is duplicated for QR Codes: {qr_codes_str}. This field must have a unique value."
                    flash(error_message, "danger")
                    return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

                other_assets_to_merge = df_to_export[other_assets_mask].copy()
                other_assets_to_merge['Asset Group'] = other_assets_to_merge['Asset Group'].str.strip()
                df_asset_group['Name'] = df_asset_group['Name'].str.strip()

                merged_others = pd.merge(
                    other_assets_to_merge, 
                    df_asset_group, 
                    left_on='Asset Group', 
                    right_on='Name', 
                    how='left'
                )
                merged_others['Asset Group'] = merged_others['Full Classification'].fillna(merged_others['Asset Group'])
                df_to_export.loc[other_assets_mask, 'Asset Group'] = merged_others['Asset Group']

        
        building_label = _get_building_label_for_filename(df_to_export)
        date_str = datetime.now().strftime("%m_%d_%Y")
        
        sdi_control_ids = df_to_export["id_print_out"].dropna().unique()
        sdi_control_label = ""
        if len(sdi_control_ids) == 1:
            sdi_control_label = f"{_safe_filename(sdi_control_ids[0])}_"
        elif len(sdi_control_ids) > 1:
            sdi_control_label = "MULTI-Package_"

        output_filename = f"SDI_Process_{sdi_control_label}{date_str}_{building_label}.xlsx"

        df2 = df_to_export.rename(columns=COLUMN_RENAME_MAP)
        for name, value in CONST_COLS.items():
            df2[name] = value

        if 'Voltage Rating' in df2.columns:
            df2['Voltage Rating (UoM)'] = ''
            condition = pd.notna(df2['Voltage Rating']) & (df2['Voltage Rating'].astype(str).str.strip() != '')
            df2.loc[condition, 'Voltage Rating (UoM)'] = 'V'

        if 'Amperage Rating' in df2.columns:
            df2['Amperage Rating (UoM)'] = ''
            condition = pd.notna(df2['Amperage Rating']) & (df2['Amperage Rating'].astype(str).str.strip() != '')
            df2.loc[condition, 'Amperage Rating (UoM)'] = 'A'

        def format_year_to_date(year_str):
            if not year_str or pd.isna(year_str):
                return year_str
            s = str(year_str).strip()
            if s.endswith('.0'):
                s = s[:-2]
            
            if s.isdigit():
                year_val = int(s)
                full_year = None
                if len(s) == 4 and 1900 < year_val < 2100:
                    full_year = year_val
                elif len(s) == 2:
                    current_year_short = datetime.now().year % 100
                    if year_val > current_year_short:
                        full_year = 1900 + year_val
                    else:
                        full_year = 2000 + year_val
                
                if full_year:
                    return f"{full_year}-01-01"
            return year_str

        if 'Date Of Manufacture Or Construction' in df2.columns:
            df2['Date Of Manufacture Or Construction'] = df2['Date Of Manufacture Or Construction'].apply(format_year_to_date)
        
        if not os.path.exists(TEMPLATE_PATH):
            raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")
        
        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active
        
        header_row, start_row = 9, 10
        norm_df_cols = {_normalize_name(c): c for c in df2.columns}
        mapping: Dict[int, str] = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=header_row, column=col_idx).value
            norm_header = _normalize_name(header_val)
            if norm_header in norm_df_cols:
                mapping[col_idx] = norm_df_cols[norm_header]
        
        if not mapping:
            raise ValueError("No template headers matched the data columns.")
        
        df_group = df2.reset_index(drop=True)
        for r, (_, row) in enumerate(df_group.iterrows(), start=start_row):
            for col_idx, df_col in mapping.items():
                val = row.get(df_col)
                ws.cell(row=r, column=col_idx, value=(None if pd.isna(val) else val))
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        _check_db_writable(DB_PATH)
        with sqlite3.connect(DB_PATH, timeout=15) as conn:
            cur = conn.cursor()
            codes_to_update = df_to_export["QR Code"].tolist()
            if codes_to_update:
                placeholders = ','.join('?' for _ in codes_to_update)
                cur.execute(f'UPDATE sdi_print_out SET print_out = 1 WHERE "QR Code" IN ({placeholders})', codes_to_update)
                conn.commit()

        return send_file(
            buffer,
            as_attachment=True,
            download_name=output_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"[ERROR] in export_to_planon: {repr(e)}")
        flash(f"⚠️ An unexpected error occurred: {str(e)}", "danger")
        return redirect(url_for("dashboard", building_code=building_code, _anchor=active_tab_anchor))

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8_003, debug=True)